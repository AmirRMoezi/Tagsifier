############################################################
#    Project: Webpages Classifier Tool                     #
#    Author: A.R. Moezi                                    #
#    Date: 23/20/11                                        #
#    Version: 1                                            #
#    Description: Classifies Pages Based on Html Tags      #
############################################################

# # Importing necessary libraries
import openpyxl as opx
import numpy as np
from urllib.request import urlopen, Request
from urllib.parse import quote
import urllib
from bs4 import BeautifulSoup
from openpyxl import Workbook as wbook
import re
import math
from sklearn.cluster import KMeans
import os
import shutil
import gzip
import argparse

parser = argparse.ArgumentParser(description="Classify webpages based on HTML structure")
parser.add_argument('--clusters', type=int, default=4, help='Number of clusters to form')
parser.add_argument('--row_start', type=int, default=1, help='Starting Excel row')
parser.add_argument('--row_end', type=int, default=70, help='Ending Excel row')
args = parser.parse_args()
path2 = str(os.path.dirname(os.path.realpath(__file__))) + "\\" 

# # Taking control Inputs from user
window_size = 5       # size of sliding window for Tags
NofClusters= args.clusters       # Number of output clusters
InputRowStart = args.row_start   # row number of Input.slsx to be started from 
InputRowEnd = args.row_end     # row number of Input.slsx to finish at

# # Importing test data
print("Importing test data ...")
if (os.path.exists(f"{path2}Input.xlsx")):
    wb = opx.load_workbook(f"{path2}Input.xlsx")
    Sheet1 = wb["Sheet1"]
    urldatabase=[]
    for i in range(InputRowStart,InputRowEnd):
        urldatabase.append(Sheet1.cell(row=i,column=1).value)
else:
    print("Input file does not exist")
print("Done")

# # Getting input raw Tags
print(f"Getting input raw text from url {InputRowStart} to {InputRowEnd} of Input.xlsx ...")
rawtextdatabase=[]
errornum=0
for index in range(0,len(urldatabase)):
    print(f"--- url number {index} loaded")
    rawtextdatabase.append([])
    try:
        url = urldatabase[index]
        url = urllib.parse.quote(url, safe=":/")
        page = urlopen((url))
        html_bytes = page.read()
        html = html_bytes.decode("utf-8")   
        soup = BeautifulSoup(html, "html.parser")
        pt = soup.find_all()
        rawtextdatabase[-1] = [x.name for x in pt]
    except Exception as e:
        #print(e)
        #print(index)
        #print(url)
        rawtextdatabase[-1] = ["##ERROR##"]
        errornum += 1
        continue

# # Sliding Window
print("Creating Dictionary ...")
pageindexdict=[]
dict=[]
inputindict=[]
indexdict = 0
for item in rawtextdatabase:
    inputindict.append([])
    item = int((window_size-1)/2) * ['*'] +  item + int((window_size-1)/2) * ['*']
    for index in range(int((window_size-1)/2), len(item) - int((window_size-1)/2)):
        window = item[index-int((window_size-1)/2) : index + int((window_size-1)/2) + 1]
        window_text=""
        for w in window:
            window_text += str(w)
        if (window_text not in dict):
            dict.append(window_text)
            indexdict = len(dict)-1
        else:
            indexdict = dict.index(window_text)
        while (indexdict>len(inputindict[-1])-1):
            inputindict[-1].append(0)
        inputindict[-1][indexdict] += 1
for id in range(0,len(inputindict)):
    while (len(inputindict[id]) < len(dict)):
        inputindict[id].append(0)
print("Done")

# # K-means Clustering ( for unknown categories )
print("Clustering ...")
sliced_inputindict = []

for item in inputindict:
    sliced_inputindict.append(item)
NofExec = 200
IterPerExec = 300
sliced_inputindict = np.array(sliced_inputindict)
kmeans = KMeans(init="random",n_clusters=NofClusters,n_init=NofExec,max_iter=IterPerExec,random_state=42)
out = kmeans.fit(sliced_inputindict)

# # Save Results to XLSX Files
if (os.path.exists(f"{path2}Results")):
    shutil.rmtree(f"{path2}Results")
os.mkdir(f"{path2}Results")
for i in range(0,NofClusters):
    indexnew=1
    book = wbook()
    path = f"{path2}Results\\{i}.xlsx"
    book.save(path)
    wbnew = opx.load_workbook(path)
    Sheetnew = wbnew["Sheet"]
    for item in range(0,len(kmeans.labels_)):
        if kmeans.labels_[item]==i:
            Sheetnew.cell(row=indexnew,column=1).value = urldatabase[item]
            indexnew += 1
    wbnew.save(path)
print("Done. outputs are shown in 'Results' directory")    
